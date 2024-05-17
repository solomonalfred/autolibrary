from db.library_manager import LibraryManager
import openpyxl


def parse_excel(file_path: str):
    wb = openpyxl.load_workbook(file_path)
    data = {}

    for sheet in wb.sheetnames:
        sheet_data = []
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, author, source, *_ = row
            if name and author and source:
                sheet_data.append({
                    "name": name,
                    "author": author,
                    "source": source
                })
        data[sheet] = sheet_data

    return data


def fill_from_excel():
    manager = LibraryManager()
    data = parse_excel("crutch/Библиотека архив.xlsx")

    for subject, books in data.items():
        for book in books:
            manager.add_record(book['name'], book['author'], subject, book['source'])


if __name__ == "__main__":
    file_path = "crutch/Библиотека архив.xlsx"
    fill_from_excel()
